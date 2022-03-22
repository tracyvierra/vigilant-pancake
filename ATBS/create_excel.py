# Author: Tracy Vierra
# Date Created: 3/21/2022
# Date Modified: 3/21/2022

# Description:

# Usage:

import openpyxl
import os

wb = openpyxl.Workbook()
wb.get_sheet_names()
wb

wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hello world!'
sheet['A2'] = 'GoodBye world!'
sheet['A3'] = 42

os.chdir('C:\\Users\\tracy\Documents\\Python Scripts\\git\\vigilant-pancake\\ATBS')
wb.save('test_create_example.xlsx')

wb.get_sheet_names
sheet2 = wb.create_sheet(index=1, title='Sheet 2')
wb.get_sheet_names

sheet2['A1'] = 'novel 1 name'
sheet2['B1'] = 'novel 2 name'
sheet2['C1'] = 'novel 3 name'

sheet2['A2'] = 'novel 1 theme'
sheet2['B2'] = 'novel 2 theme'
sheet2['C2'] = 'novel 3 theme'

sheet2['A3'] = 'novel 1 schedule'
sheet2['B3'] = 'novel 2 schedule'
sheet2['C3'] = 'novel 3 schedule'

sheet2.title = 'Sheet 2A'
wb.get_sheet_names()

wb.save('test_create_example_2.xlsx')


sheet3 = wb.create_sheet(index=2, title='Sheet 3')
sheets = wb.get_sheet_names()
print(sheets)

sheet3['A1'] = 'adventure 1 name'
sheet3['B1'] = 'adventure 2 name'
sheet3['C1'] = 'adventure 3 name'

sheet3['A2'] = 'adventure 1 theme'
sheet3['B2'] = 'adventure 2 theme'
sheet3['C2'] = 'adventure 3 theme'

sheet3['A3'] = 'adventure 1 schedule'
sheet3['B3'] = 'adventure 2 schedule'
sheet3['C3'] = 'adventure 3 schedule'

wb.save('test_create_example_3.xlsx')

